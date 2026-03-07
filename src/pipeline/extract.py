"""Extract data from Excel files."""

import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl


def _serialize_cell(value: Any) -> Any:
    """Convert a cell value to a JSON-serialisable type."""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def extract_excel(file_path: str) -> dict[str, list[dict]]:
    """Parse an xlsm/xlsx file and return all sheet data as a dict of sheet_name → rows.

    Each row is a dict mapping column headers (row 0) to cell values.
    All values are JSON-serialisable.  Formulas are resolved to their cached
    values via ``data_only=True``.
    """
    path = Path(file_path)
    wb = openpyxl.load_workbook(path, data_only=True)

    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            result[sheet_name] = []
            continue

        # Build header list – fall back to positional names for blank/duplicate cells
        seen: dict[str, int] = {}
        headers: list[str] = []
        for i, h in enumerate(raw_headers):
            name = str(h).strip() if h is not None else f"col_{i}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            headers.append(name)

        sheet_rows: list[dict] = []
        for row in rows_iter:
            serialised = {headers[i]: _serialize_cell(v) for i, v in enumerate(row)}
            sheet_rows.append(serialised)

        result[sheet_name] = sheet_rows

    wb.close()
    return result
